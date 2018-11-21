import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

def bubbleSort(arr,arr2):
    n = len(arr)
    for i in range(0,n):
        for j in range(0, n-i-1):
            if arr[j] < arr[j+1] :
                arr[j], arr[j+1] = arr[j+1], arr[j]
                arr2[j], arr2[j+1] = arr2[j+1], arr2[j]

ws = openpyxl.load_workbook('MI_score.xlsx')
w2 = ws.get_sheet_names()
sheet = ws[w2[1]]

ws1 = openpyxl.load_workbook('features.xlsx')
w3 = ws1.get_sheet_names()
sheet1 = ws1[w3[0]]

from openpyxl import Workbook
wb = Workbook()

#target = wb.copy_worksheet(sheet)

row_count = sheet.max_row
column_count = sheet.max_column

print(row_count)
print(column_count)
for i in range(0,18):
    m = i+1
    index =[]
    for q in range(1,column_count):
        index.append(q)
    listx = []
    for j in range(1,column_count):
        listx.append(sheet.cell(row=m,column=j+1).value)
    bubbleSort(listx,index)


    str = sheet.cell(row=m,column=1).value
    wx = wb.create_sheet(str)
    

    for k in range(1,20):
        wx.cell(row = k+1,column=1,value = sheet1.cell(row=k+1,column=1).value)


    list =  []
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    list6 = []
    for p in range(1,21):
        list.append(sheet1.cell(row=p,column=index[0]+1).value)
        list1.append(sheet1.cell(row=p,column=index[1]+1).value)
        list2.append(sheet1.cell(row=p,column=index[2]+1).value)
        list3.append(sheet1.cell(row=p,column=index[3]+1).value)
        list4.append(sheet1.cell(row=p,column=index[4]+1).value)
        list5.append(sheet1.cell(row=p,column=index[5]+1).value)
        list6.append(sheet1.cell(row=p,column=index[6]+1).value)



    for j in range(0,20):
        wx.cell(row=j+1,column=2,value=list[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=3,value=list1[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=4,value=list2[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=5,value=list3[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=6,value=list4[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=7,value=list5[j])
    
    for j in range(0,20):
        wx.cell(row=j+1,column=8,value=list6[j])
            
    
wb.save('Top_seven_features.xlsx')



















