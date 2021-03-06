import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
from sklearn.model_selection import train_test_split


xx = pd.read_excel('Top_seven_features.xlsx',sheet_name = 'Total')
x = xx.iloc[:,0:]
ym = pd.read_excel('turker_scores.xlsx')
y = ym.iloc[0:,18:19]

print(y.shape)

x_tr,x_ts,y_tr,y_ts = train_test_split(x,y,test_size=0.1)


from sklearn.preprocessing import StandardScaler
sc_X = StandardScaler()
sc_y = StandardScaler()
x_tr = sc_X.fit_transform(x_tr)
x_ts = sc_X.transform(x_ts)
y_tr = sc_y.fit_transform(y_tr)
y_ts = sc_y.fit_transform(y_ts)


from sklearn.svm import SVR
regressor = SVR(kernel='rbf')
regressor.fit(x_tr,y_tr)


y_pred = regressor.predict(x_ts)
regressor.score(x_tr,y_tr)
regressor.score(x_ts,y_ts)


from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator = regressor, X = x_tr, y = y_tr, cv = 5,scoring='neg_mean_squared_error')
accuracies = accuracies* (-1)

accuracies.mean()
accuracies.std()



from sklearn.ensemble import RandomForestRegressor
regressor1 = RandomForestRegressor(n_estimators = 100, random_state = 0)
regressor1.fit(x_tr,y_tr)


y_pred = regressor.predict(x_ts)
regressor1.score(x_tr,y_tr)
regressor1.score(x_ts,y_ts)


from sklearn.model_selection import cross_val_score

accuracies1 = cross_val_score(estimator = regressor1, X = x_tr, y = y_tr, cv = 5)
accuracies1.mean()
accuracies1.std()


import openpyxl
ws = openpyxl.load_workbook('Top_seven_features.xlsx')
#w2 = ws.get_sheet_names()
#sheet = ws[w2[0]]
sheet = ws['Total']
sheet.cell(row=21,column=1,value = 'SVR')
sheet.cell(row=22,column=1,value = 'SVR-->5 values of accuracy by K fold CV')
for i in range(2,7):
    sheet.cell(row=22,column=i,value =accuracies[i-2])
sheet.cell(row=23,column=1,value = 'MeanAccuracies')
sheet.cell(row=23,column=2,value =accuracies.mean())
sheet.cell(row=24,column=1,value = 'StdAccuracies')
sheet.cell(row=24,column=2,value =accuracies.std())
sheet.cell(row=26,column=1,value = 'Random Forest')
sheet.cell(row=27,column=1,value = 'RForest-->5 values of accuracy by K fold CV')
for j in range(2,7):
    sheet.cell(row=27,column=j,value =accuracies1[j-2])
sheet.cell(row=28,column=1,value = 'MeanAccuracies')
sheet.cell(row=28,column=2,value =accuracies1.mean())
sheet.cell(row=29,column=1,value = 'StdAccuracies')
sheet.cell(row=29,column=2,value =accuracies1.std())

plt.plot(accuracies)
pl.title('Total Accuracies')
plt.show()
ws.save('Top_seven_features.xlsx')
